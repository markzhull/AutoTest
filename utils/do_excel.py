from openpyxl import load_workbook

class DoExcel():
  def __init__(self,file_path,sheet_name):
    self.file_path = file_path
    self.sheet_name = sheet_name

  def do_excel(self):
    wb = load_workbook(self.file_path)
    sheet = wb[self.sheet_name]
    test_data = []
    for i in range(2,sheet.max_row+1):
      sub_data={}
      sub_data['case_id'] = sheet.cell(i,1).value
      sub_data['title'] = sheet.cell(i,2).value
      sub_data['method'] = sheet.cell(i,3).value
      sub_data['url'] = sheet.cell(i,4).value
      sub_data['data'] = sheet.cell(i,5).value
      sub_data['ExpectedResult'] = sheet.cell(i,6).value

      test_data.append(sub_data)
    return test_data

  def write_data(self,r,c,ActualResult):
    wb = load_workbook(self.file_path)
    sheet = wb[self.sheet_name]
    sheet.cell(r,c).value = ActualResult
    wb.save(self.file_path)


if __name__ == '__main__':
    res = DoExcel('/Users/zhuhongcheng/PycharmProjects/MyAutoTestTool/data/credit_card/信用卡分销接口测试用例.xlsx', 'test_data').do_excel()
    print(res)
